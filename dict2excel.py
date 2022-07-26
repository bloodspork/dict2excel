# author: bloodspork
# date  : 20220726

from openpyxl import Workbook

def autoFitColumnWidth(ws):
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = \
                    max((dims.get(cell.column_letter, 0), len(str(cell.value))))
    for col, value in dims.items():
        ws.column_dimensions[col].width = value + 4


def dict2excel(dct: dict, filepath: str):
    wb = Workbook()
    for sheetname in dct:
        sheet = wb.create_sheet(sheetname)
        row_list = dct[sheetname]

        colnum = 1
        rownum = 1
        for row in row_list:
            if rownum == 1:
                for fieldname in row:
                    sheet.cell(rownum, colnum, fieldname)
                    colnum += 1
                rownum += 1

            colnum = 1
            for fieldname in row:
                sheet.cell(rownum, colnum, row[fieldname])
                colnum += 1
            rownum += 1
        autoFitColumnWidth(sheet)

    del wb['Sheet']
    wb.save(filepath)


